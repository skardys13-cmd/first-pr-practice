#!/usr/bin/env python3
"""Generates the workbook from the SAME inputs.json the page is built from.
Every figure is a live formula; nothing is a pasted result."""
import json, datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chart import LineChart, BarChart, Reference

I = json.load(open("inputs.json"))
AS_OF = dt.date.fromisoformat(I["meta"]["as_of"])
HORIZON_MONTHS = 72
TIERS = [b["tier"] for b in I["book"]]

# ---------------------------------------------------------------- styling
FONT = "Arial"
INK = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
TITLE = Font(name=FONT, size=14, bold=True)
H2 = Font(name=FONT, size=11, bold=True)
INPUT_F = Font(name=FONT, size=10, color="0000FF")               # blue  = typed input
LINK_F = Font(name=FONT, size=10, color="008000")                # green = link to another sheet
MUTED = Font(name=FONT, size=9, color="666666")
WARN_F = Font(name=FONT, size=10, color="C00000", bold=True)
IN_FILL = PatternFill("solid", fgColor="FFF7DC")                 # inputs
KEY_FILL = PatternFill("solid", fgColor="FFF2A8")                # key assumptions to fill in
FX_FILL = PatternFill("solid", fgColor="FFFFFF")                 # formulas
HDR_FILL = PatternFill("solid", fgColor="1C6B52")
HDR_F = Font(name=FONT, size=9, bold=True, color="FFFFFF")
BAND = PatternFill("solid", fgColor="F2F5F2")
thin = Side(style="thin", color="D5DAD5")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

MONEY = '$#,##0;($#,##0);-'
MONEY2 = '$#,##0.00;($#,##0.00);-'
PCT = '0.0%'
HOURS = '#,##0.0'
NUM = '#,##0'
DATE = 'mmmm yyyy'

# Count of ledger rows, computed the same way the Assumptions tab builds them.
N_INPUTS = (len(I["constants"])
            + sum(2 + (1 if m["is_operations"] else 0) for m in I["roster"])
            + len(I["book"]) + 1 + 1 + len(I["tasks"]))

wb = Workbook()
named = {}

def sheet(title):
    ws = wb.create_sheet(title); ws.sheet_view.showGridLines = False; return ws

def hdr(ws, row, labels, widths=None):
    for i, l in enumerate(labels, 1):
        c = ws.cell(row=row, column=i, value=l); c.font = HDR_F; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = BOX
    if widths:
        for i, w in enumerate(widths, 1): ws.column_dimensions[gcl(i)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)

def put(ws, r, c, v, font=INK, fmt=None, fill=None, align=None, border=True):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font
    if fmt: cell.number_format = fmt
    if fill: cell.fill = fill
    if align: cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=(align == "left"))
    if border: cell.border = BOX
    return cell

def name(nm, sheetname, ref):
    wb.defined_names[nm] = DefinedName(nm, attr_text=f"'{sheetname}'!{ref}")

TAG_DV = '"MEASURED,OBSERVED,BENCHMARK,ESTIMATED,PLACEHOLDER"'

# ================================================================= README
ws = wb.active; ws.title = "README"; ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 96
r = 2
put(ws, r, 2, "Operations capacity & economics model", TITLE, border=False); r += 1
put(ws, r, 2, f"Working model · as of {I['meta']['as_of']} · schema {I['meta']['schema_version']}", MUTED, border=False); r += 2

def block(title, lines):
    global r
    put(ws, r, 2, title, H2, border=False); r += 1
    for a, b in lines:
        put(ws, r, 2, a, BOLD if a else INK, align="left", border=False)
        put(ws, r, 3, b, INK, align="left", border=False)
        ws.row_dimensions[r].height = max(14, 13 * (1 + len(b) // 95))
        r += 1
    r += 1

block("What this is", [
    ("", "A model of how much operations work the firm carries, who carries it, what it costs, and when it runs out of room."),
    ("", "Built by the paraplanner (role R4) as a working tool, not a firm document. It is the same engine as the interactive page — both are generated from one inputs file, so they cannot disagree."),
    ("Who to correct it", "Everything in here is meant to be argued with. Change any blue or yellow cell and the whole workbook recalculates."),
])
block("The colour key", [
    ("Blue text on cream", "A typed input. This is a number somebody chose. Change these."),
    ("Blue text on yellow", "A key assumption or a placeholder that still needs a real value. Change these first."),
    ("Green text", "A link to a value on another sheet. Do not overwrite."),
    ("Black text, white", "A formula. Do not overwrite — it will stop recalculating."),
    ("Red text", "A figure that is currently built on a placeholder, or a result that needs attention."),
])
block("What is measured and what is not", [
    ("", f"Every input carries a provenance tag. Of {0} inputs, most are still PLACEHOLDER — numbers invented so the model would run end to end. See the Assumptions tab; the tag column is the whole point of this workbook."),
    ("MEASURED", I["tag_meta"]["MEASURED"]["desc"] + " Strongest evidence in the model."),
    ("OBSERVED", I["tag_meta"]["OBSERVED"]["desc"]),
    ("BENCHMARK", I["tag_meta"]["BENCHMARK"]["desc"]),
    ("ESTIMATED", I["tag_meta"]["ESTIMATED"]["desc"]),
    ("PLACEHOLDER", I["tag_meta"]["PLACEHOLDER"]["desc"] + " Treat any figure resting on one as a structure, not a finding."),
])
block("How to change the fee schedule", [
    ("1", "Go to the FeeSchedule tab. Each row is one band: the amount it starts at, the amount it ends at, and the rate charged on the portion of assets inside it."),
    ("2", "Edit From / To / Rate, or add a row. If you add one, extend the named range Fee_From, Fee_To, Fee_Rate and Fee_Delta to include it."),
    ("3", "The schedule is TIERED (marginal): each band charges its own rate on its own slice, not one rate on the whole balance."),
    ("4", "The tab carries three hand-worked test households. If those three stop matching the hand column, the schedule has been broken — most likely by turning it into a flat rate by accident."),
])
block("How to switch scenario", [
    ("Assumptions!C5", "One cell drives everything. 1 = Today · 2 = She's gone, nothing changes · 3 = She's gone, I own operations."),
    ("", "Every case tab reads that cell. Nothing else needs changing to move between the three."),
])
block("Rules this model runs under", [
    ("No client data", "Counts, tiers, ranges and averages only. No names, no account numbers, no balances, no statements. The smallest unit anywhere in it is a tier."),
    ("Not a compliance opinion", "Compliance ownership stays with the firm's designated person. Where this counts minutes on a compliance task it is counting minutes, not judging adequacy."),
    ("No credentials", "No password, login or vendor credential appears anywhere in this workbook, and none should ever be added."),
    ("Labels", "Role labels come from the Labels tab. Fill in the 'Real' column there for internal use; the published version of this model uses the anonymous column only."),
])
put(ws, r, 2, "Tab order", H2, border=False); r += 1
for nm, desc in [
    ("Assumptions", "The ledger. Every input, its value, its tag, where it came from, who could confirm it, when it was last checked."),
    ("Labels", "Role labels. The single place names live."),
    ("Roster", "Input: who is here, at what FTE, on what compensation, and how much of their time is operations."),
    ("Book", "Input: households and average AUM by tier. Never per household."),
    ("FeeSchedule", "Input: the tiered fee bands, plus three hand-worked test households."),
    ("Tasks", "Input: the operations task catalogue. Minutes and occurrences. This is the heart of the model."),
    ("Engine", "All calculations. No typed numbers anywhere on this sheet."),
    ("Case1_Departure", "What happens when the consultant leaves."),
    ("Case2_Capacity", "Where capacity breaks — month by month, and the crossing."),
    ("Case3_CostToServe", "What each tier costs to serve, in layers."),
    ("Case4_Seat", "The operations seat: cost, what it protects, what it displaces."),
    ("Dashboard", "The summary, with charts."),
]:
    put(ws, r, 2, nm, BOLD, align="left", border=False)
    put(ws, r, 3, desc, INK, align="left", border=False); r += 1

# ================================================================= LABELS
ls = sheet("Labels")
hdr(ls, 1, ["Key", "Anonymous label", "Real label (fill in for internal use only)"], [14, 34, 46])
lr = 2
for k, v in I["labels"].items():
    if k == "_note": continue
    put(ls, lr, 1, k, INK); put(ls, lr, 2, v["anon"], INPUT_F, fill=IN_FILL)
    put(ls, lr, 3, v["real"], INPUT_F, fill=KEY_FILL); lr += 1
put(ls, lr + 1, 1, "ANONYMISE", BOLD)
put(ls, lr + 1, 2, 1, INPUT_F, fmt=NUM, fill=KEY_FILL)
put(ls, lr + 1, 3, "1 = use the anonymous column everywhere · 0 = use the real column. Ships as 1.", MUTED)
name("Anonymise", "Labels", f"$B${lr+1}")
name("Label_Key", "Labels", f"$A$2:$A${lr-1}")
name("Label_Anon", "Labels", f"$B$2:$B${lr-1}")
name("Label_Real", "Labels", f"$C$2:$C${lr-1}")
LBL = lambda key: f'INDEX(IF(Anonymise=1,Label_Anon,Label_Real),MATCH("{key}",Label_Key,0))'
# INDEX(IF(...)) needs array evaluation; use a plain switch instead:
LBL = lambda key: f'IF(Anonymise=1,INDEX(Label_Anon,MATCH("{key}",Label_Key,0)),INDEX(Label_Real,MATCH("{key}",Label_Key,0)))'
print("labels rows", lr - 2)

# ================================================================= ASSUMPTIONS
asm = sheet("Assumptions")
for i, w in enumerate([9, 42, 17, 12, 14, 60, 40, 13], 1): asm.column_dimensions[gcl(i)].width = w
put(asm, 1, 1, "ASSUMPTION LEDGER", TITLE, border=False)
put(asm, 2, 1, "Every input in the model. If a number anywhere else is wrong, it is wrong here first.", MUTED, border=False)
put(asm, 4, 1, "SCENARIO SWITCH — this one cell drives every case tab", H2, border=False)

CTRL = [("Scenario", 1, NUM, "1 = Today  ·  2 = She's gone, nothing changes  ·  3 = She's gone, I own operations", "Scenario", True),
        ("Scenario name", '=IF(Scenario=1,"Today",IF(Scenario=2,"She\'s gone, nothing changes","She\'s gone, I own operations"))', None, "Reads the switch above.", None, False),
        ("Consultant on the roster", "=IF(Scenario=1,1,0)", NUM, "1 = still here. Drives the Roster active flag.", "Consultant_Present", False),
        ("I take the operations seat", "=IF(Scenario=3,1,0)", NUM, "1 = my row moves to the operations allocation and compensation below.", "Resp_Internal", False),
        ("Minutes per task (efficiency)", 1.0, PCT, "Global multiplier on every task timing. 90% is a 10% process improvement.", "Efficiency", True),
        ("Extra operations FTE", 0.0, '#,##0.0', "Additive only in the workbook. Use it to test a hire.", "Extra_Ops_Fte", True),
        ("My operations share if I take the seat", 0.85, PCT, "Case 4. The rest of my week stays paraplanning.", "Seat_Ops_Alloc", True),
        ("As of", AS_OF, 'yyyy-mm-dd', "Every projected date counts forward from here.", "As_Of", True),
        ("Projection horizon (months)", HORIZON_MONTHS, NUM, "Length of the capacity projection on Case2.", "Horizon", True)]
r = 5
for label, val, fmt, note, nm, is_in in CTRL:
    put(asm, r, 2, label, BOLD)
    c = put(asm, r, 3, val, INPUT_F if is_in else INK, fmt=fmt, fill=KEY_FILL if is_in else FX_FILL)
    put(asm, r, 6, note, MUTED, align="left")
    if nm: name(nm, "Assumptions", f"$C${r}")
    r += 1
dv = DataValidation(type="whole", operator="between", formula1=1, formula2=3, allow_blank=False,
                    error="Scenario must be 1, 2 or 3.", errorTitle="Scenario")
asm.add_data_validation(dv); dv.add(asm["C5"])

LEDGER_HDR = r + 1
hdr(asm, LEDGER_HDR, ["Ref", "Input", "Value", "Unit", "Tag", "Where it came from", "Who could confirm it", "Last checked"])
asm.freeze_panes = asm.cell(row=LEDGER_HDR + 1, column=1)
tag_dv = DataValidation(type="list", formula1=TAG_DV, allow_blank=False,
                        error="Provenance cannot be blank. Pick one of the five tags.", errorTitle="Provenance required")
asm.add_data_validation(tag_dv)

NAME_MAP = {"benefits_load": "Benefits_Load", "productive_hours_year": "Productive_Hours",
    "growth_rate_households": "Growth_Rate", "switching_uplift": "Switching_Uplift",
    "firm_overhead_per_household": "Firm_Overhead_PH", "fee_realisation": "Fee_Realisation",
    "consultant_departure_days": "Departure_Days", "ops_seat_comp": "Ops_Seat_Comp",
    "backfill_consultant_cost": "Backfill_Cost", "service_failure_cost_per_event": "Failure_Cost",
    "service_failures_avoided": "Failures_Avoided", "paraplanning_backfill_comp": "Para_Backfill_Comp",
    "advisor_hour_basis": "Advisor_Basis"}
UNIT_FMT = {"usd": MONEY, "pct": PCT, "hours": NUM, "days": NUM, "count": NUM, "fte": '#,##0.00',
            "minutes": NUM, "households": NUM, "choice": None, "schedule": None, "mix": None}

lr = LEDGER_HDR + 1
ledger_rows = {}
n = 0
for k, c in I["constants"].items():
    n += 1; ref = f"C-{n}"
    put(asm, lr, 1, ref, MUTED); put(asm, lr, 2, c["name"], INK, align="left")
    is_ph = c["tag"] == "PLACEHOLDER"
    put(asm, lr, 3, c["v"], INPUT_F, fmt=UNIT_FMT.get(c["unit"]), fill=KEY_FILL if is_ph else IN_FILL)
    put(asm, lr, 4, c["unit"], MUTED)
    put(asm, lr, 5, c["tag"], WARN_F if is_ph else INK); tag_dv.add(asm.cell(row=lr, column=5))
    put(asm, lr, 6, c["src"], MUTED, align="left"); put(asm, lr, 7, c["who"], MUTED, align="left")
    put(asm, lr, 8, c["checked"], WARN_F if c["checked"] == "never" else MUTED)
    if k in NAME_MAP: name(NAME_MAP[k], "Assumptions", f"$C${lr}")
    ledger_rows[k] = lr; lr += 1
CONST_LAST = lr - 1

# ================================================================= ROSTER
rs = sheet("Roster")
hdr(rs, 1, ["Role", "Label", "FTE", "Advisor?", "Operations?", "Ops share", "Compensation", "Comp tag",
            "Contractor?", "End date", "Active", "Effective ops share", "Effective comp",
            "Loaded cost", "Loaded $/hour", "Ops hours available"],
    [8, 30, 8, 9, 11, 11, 14, 13, 11, 12, 8, 15, 14, 13, 13, 14])
rtag_dv = DataValidation(type="list", formula1=TAG_DV, allow_blank=False,
                         error="Provenance cannot be blank.", errorTitle="Provenance required")
rs.add_data_validation(rtag_dv)
roster_rows = {}
rr = 2
for m in I["roster"]:
    rid = m["role_id"]
    put(rs, rr, 1, rid, INK)
    put(rs, rr, 2, f"={LBL(rid)}", LINK_F, align="left")
    put(rs, rr, 3, m["fte"], INPUT_F, fmt='#,##0.00', fill=IN_FILL)
    put(rs, rr, 4, 1 if m["is_advisor"] else 0, INPUT_F, fmt=NUM, fill=IN_FILL)
    put(rs, rr, 5, 1 if m["is_operations"] else 0, INPUT_F, fmt=NUM, fill=IN_FILL)
    put(rs, rr, 6, m["ops_allocation"], INPUT_F, fmt=PCT, fill=IN_FILL)
    is_ph = m["comp_tag"] == "PLACEHOLDER"
    put(rs, rr, 7, m["comp"], INPUT_F, fmt=MONEY, fill=KEY_FILL if is_ph else IN_FILL)
    put(rs, rr, 8, m["comp_tag"], WARN_F if is_ph else INK); rtag_dv.add(rs.cell(row=rr, column=8))
    put(rs, rr, 9, 1 if m["is_contractor"] else 0, INPUT_F, fmt=NUM, fill=IN_FILL)
    put(rs, rr, 10, m["end_date"] or "", INPUT_F, fill=IN_FILL)
    put(rs, rr, 11, f'=IF(A{rr}="R7",Consultant_Present,1)', INK, fmt=NUM)
    put(rs, rr, 12, f'=IF(AND(A{rr}="R4",Resp_Internal=1),Seat_Ops_Alloc,F{rr})', INK, fmt=PCT)
    put(rs, rr, 13, f'=IF(AND(A{rr}="R4",Resp_Internal=1),Ops_Seat_Comp,G{rr})', INK, fmt=MONEY)
    put(rs, rr, 14, f'=M{rr}*(1+IF(I{rr}=1,0,Benefits_Load))', INK, fmt=MONEY)
    put(rs, rr, 15, f'=IF(C{rr}*Productive_Hours=0,0,N{rr}/(C{rr}*Productive_Hours))', INK, fmt=MONEY2)
    put(rs, rr, 16, f'=K{rr}*E{rr}*C{rr}*L{rr}*Productive_Hours', INK, fmt=HOURS)
    roster_rows[rid] = rr; rr += 1
RN = rr - 1
put(rs, rr + 1, 2, "Total FTE", BOLD); put(rs, rr + 1, 3, f"=SUMPRODUCT(K2:K{RN},C2:C{RN})", BOLD, fmt='#,##0.00')
put(rs, rr + 1, 15, "Operations hours available", BOLD)
put(rs, rr + 1, 16, f"=SUM(P2:P{RN})+Extra_Ops_Fte*Productive_Hours", BOLD, fmt=HOURS)
put(rs, rr + 3, 2, "Loaded $/hour divides by the role's OWN hours (FTE x productive hours), so a fractional "
                   "role is not costed as though it were full time. Contractors carry no benefits load.", MUTED, align="left")
name("Roster_Active", "Roster", f"$K$2:$K${RN}")
name("Roster_Adv", "Roster", f"$D$2:$D${RN}")
name("Roster_FTE", "Roster", f"$C$2:$C${RN}")
name("Roster_Loaded", "Roster", f"$N$2:$N${RN}")
name("Roster_Hourly", "Roster", f"$O$2:$O${RN}")
name("Roster_OpsHours", "Roster", f"$P$2:$P${RN}")
name("Ops_Hours_Available", "Roster", f"$P${rr+1}")
name("Total_FTE", "Roster", f"$C${rr+1}")

# ================================================================= BOOK
bk = sheet("Book")
hdr(bk, 1, ["Tier", "Label", "Households", "Average AUM", "Accounts per household", "New-business mix",
            "Tag", "Total AUM", "Accounts in tier", "New households / yr", "Fee per household",
            "Revenue per household", "Revenue from tier"],
    [7, 16, 12, 14, 15, 14, 14, 15, 13, 14, 13, 14, 15])
btag_dv = DataValidation(type="list", formula1=TAG_DV, allow_blank=False, errorTitle="Provenance required",
                         error="Provenance cannot be blank.")
bk.add_data_validation(btag_dv)
br = 2
for b in I["book"]:
    put(bk, br, 1, b["tier"], BOLD); put(bk, br, 2, b["label"], INK)
    put(bk, br, 3, b["households"], INPUT_F, fmt=NUM, fill=IN_FILL)
    put(bk, br, 4, b["avg_aum"], INPUT_F, fmt=MONEY, fill=IN_FILL)
    put(bk, br, 5, b["accounts_per_household"], INPUT_F, fmt='#,##0.0', fill=IN_FILL)
    put(bk, br, 6, b["new_mix"], INPUT_F, fmt=PCT, fill=IN_FILL)
    put(bk, br, 7, b["tag"], INK); btag_dv.add(bk.cell(row=br, column=7))
    put(bk, br, 8, f"=C{br}*D{br}", INK, fmt=MONEY)
    put(bk, br, 9, f"=C{br}*E{br}", INK, fmt='#,##0.0')
    put(bk, br, 10, f"=Total_HH*Growth_Rate*F{br}", INK, fmt='#,##0.0')
    put(bk, br, 11, f"=SUMPRODUCT((D{br}>Fee_From)*(D{br}-Fee_From)*Fee_Delta)", INK, fmt=MONEY)
    put(bk, br, 12, f"=K{br}*Fee_Realisation", INK, fmt=MONEY)
    put(bk, br, 13, f"=C{br}*L{br}", INK, fmt=MONEY)
    br += 1
BN = br - 1
put(bk, br, 2, "Whole book", BOLD)
put(bk, br, 3, f"=SUM(C2:C{BN})", BOLD, fmt=NUM)
put(bk, br, 8, f"=SUM(H2:H{BN})", BOLD, fmt=MONEY)
put(bk, br, 10, f"=SUM(J2:J{BN})", BOLD, fmt='#,##0.0')
put(bk, br, 13, f"=SUM(M2:M{BN})", BOLD, fmt=MONEY)
put(bk, br + 2, 2, "Households and average AUM are the inputs; total AUM is a formula. Never put a single "
                   "household on this sheet — if a figure would identify one, it is at the wrong grain.", MUTED, align="left")
put(bk, br + 3, 2, "New-business mix must sum to 100%.", MUTED, align="left")
put(bk, br + 3, 6, f"=SUM(F2:F{BN})", BOLD, fmt=PCT)
name("Total_HH", "Book", f"$C${br}")
name("Book_HH", "Book", f"$C$2:$C${BN}")
name("Book_AvgAum", "Book", f"$D$2:$D${BN}")
name("Book_Acct", "Book", f"$E$2:$E${BN}")
name("Book_Mix", "Book", f"$F$2:$F${BN}")
name("Book_HHAcct", "Book", f"$I$2:$I${BN}")
name("Book_New", "Book", f"$J$2:$J${BN}")
name("Book_RevPH", "Book", f"$L$2:$L${BN}")
name("Revenue_Total", "Book", f"$M${br}")
name("AUM_Total", "Book", f"$H${br}")
BOOK_FIRST = 2

# ================================================================= FEE SCHEDULE
fs = sheet("FeeSchedule")
hdr(fs, 1, ["Band", "From", "To", "Rate", "Rate step (rate less previous rate)"], [8, 15, 15, 11, 30])
OPEN_TOP = 1e15
fr = 2
for i, b in enumerate(I["fee_schedule"]["bands"]):
    put(fs, fr, 1, i + 1, INK, fmt=NUM)
    put(fs, fr, 2, b["from"], INPUT_F, fmt=MONEY, fill=IN_FILL)
    put(fs, fr, 3, b["to"] if b["to"] is not None else OPEN_TOP, INPUT_F, fmt=MONEY, fill=IN_FILL)
    put(fs, fr, 4, b["rate"], INPUT_F, fmt='0.00%', fill=IN_FILL)
    put(fs, fr, 5, f"=D{fr}" if i == 0 else f"=D{fr}-D{fr-1}", INK, fmt='0.00%')
    fr += 1
FN = fr - 1
name("Fee_From", "FeeSchedule", f"$B$2:$B${FN}")
name("Fee_To", "FeeSchedule", f"$C$2:$C${FN}")
name("Fee_Rate", "FeeSchedule", f"$D$2:$D${FN}")
name("Fee_Delta", "FeeSchedule", f"$E$2:$E${FN}")
put(fs, fr + 1, 1, "TIERED / MARGINAL", BOLD)
put(fs, fr + 1, 2, I["fee_schedule"]["basis_note"], MUTED, align="left")
put(fs, fr + 2, 2, "The fee is computed as SUMPRODUCT((AUM>Fee_From)*(AUM-Fee_From)*Fee_Delta) — the rate-step "
                   "form of a marginal schedule. It is mathematically identical to summing each band's own slice, "
                   "and it survives adding or removing bands.", MUTED, align="left")
put(fs, fr + 3, 2, "The open top band's To is 1E+15 rather than blank, so the arithmetic has an upper bound.", MUTED, align="left")

HC = fr + 5
put(fs, HC, 1, "HAND CHECK — three households, worked on paper", H2, border=False)
put(fs, HC + 1, 1, "If any row below says FAIL, the schedule has been broken. The usual cause is turning a tiered "
                   "schedule into a flat rate by accident.", MUTED, align="left")
hdr(fs, HC + 2, ["Case", "AUM", "Model says", "Hand-worked", "Check", "The working"], [26, 14, 14, 14, 10, 74])
fs.freeze_panes = "A2"
CHECKS = [("Below the first breakpoint", 145000, 1450.00, "145,000 x 1.00% = 1,450.00"),
          ("Just above a breakpoint", 520000, 5170.00,
           "500,000 x 1.00% = 5,000.00  +  20,000 x 0.85% = 170.00  =  5,170.00     (a flat 0.85% on the whole "
           "balance would give 4,420.00 — a $750 error, and the commonest mistake in these models)"),
          ("Well above", 5100000, 33600.00,
           "500,000 x 1.00% = 5,000  +  500,000 x 0.85% = 4,250  +  2,000,000 x 0.70% = 14,000  +  "
           "2,000,000 x 0.50% = 10,000  +  100,000 x 0.35% = 350  =  33,600  (effective rate 0.659%)")]
hr_ = HC + 3
for label, aum, hand, working in CHECKS:
    put(fs, hr_, 1, label, INK, align="left")
    put(fs, hr_, 2, aum, INPUT_F, fmt=MONEY, fill=IN_FILL)
    put(fs, hr_, 3, f"=SUMPRODUCT((B{hr_}>Fee_From)*(B{hr_}-Fee_From)*Fee_Delta)", INK, fmt=MONEY2)
    put(fs, hr_, 4, hand, INK, fmt=MONEY2)
    put(fs, hr_, 5, f'=IF(ROUND(C{hr_}-D{hr_},2)=0,"PASS","FAIL")', BOLD)
    put(fs, hr_, 6, working, MUTED, align="left")
    fs.row_dimensions[hr_].height = 30
    hr_ += 1
put(fs, hr_ + 1, 1, "All three pass", BOLD)
put(fs, hr_ + 1, 5, f'=IF(COUNTIF(E{HC+3}:E{hr_-1},"PASS")=3,"PASS","FAIL")', BOLD)

# ================================================================= TASKS
tk = sheet("Tasks")
hdr(tk, 1, ["ID", "Task", "Trigger basis", "Minutes per occurrence", "Occurrences / yr"] +
    [f"Applies {t}" for t in TIERS] + ["Owner", "Work type", "Tag", "Minutes / yr", "Hours / yr"],
    [7, 52, 15, 12, 12, 9, 9, 9, 9, 9, 12, 14, 12, 11])
ttag_dv = DataValidation(type="list", formula1=TAG_DV, allow_blank=False, errorTitle="Provenance required",
                         error="Provenance cannot be blank. Time it, or leave it PLACEHOLDER.")
tk.add_data_validation(ttag_dv)
basis_dv = DataValidation(type="list", formula1='"firm,household,account,new_household"', allow_blank=False)
tk.add_data_validation(basis_dv)
wt_dv = DataValidation(type="list", formula1='"operations,advisory"', allow_blank=False)
tk.add_data_validation(wt_dv)
tr = 2
task_rows = {}
for t in I["tasks"]:
    put(tk, tr, 1, t["id"], MUTED); put(tk, tr, 2, t["task"], INK, align="left")
    put(tk, tr, 3, t["basis"], INPUT_F, fill=IN_FILL); basis_dv.add(tk.cell(row=tr, column=3))
    put(tk, tr, 4, t["minutes"], INPUT_F, fmt=NUM, fill=KEY_FILL)
    put(tk, tr, 5, t["occurrences"], INPUT_F, fmt='#,##0.00', fill=KEY_FILL)
    for i, tier in enumerate(TIERS):
        put(tk, tr, 6 + i, 1 if tier in t["tiers"] else 0, INPUT_F, fmt=NUM, fill=IN_FILL)
    put(tk, tr, 10, t["owner"], INPUT_F, fill=IN_FILL)
    put(tk, tr, 11, t["work_type"], INPUT_F, fill=IN_FILL); wt_dv.add(tk.cell(row=tr, column=11))
    put(tk, tr, 12, t["tag"], WARN_F if t["tag"] == "PLACEHOLDER" else INK); ttag_dv.add(tk.cell(row=tr, column=12))
    hh_ = "+".join(f"{gcl(6+i)}{tr}*Book!$C${BOOK_FIRST+i}" for i in range(len(TIERS)))
    ac_ = "+".join(f"{gcl(6+i)}{tr}*Book!$I${BOOK_FIRST+i}" for i in range(len(TIERS)))
    nw_ = "+".join(f"{gcl(6+i)}{tr}*Book!$J${BOOK_FIRST+i}" for i in range(len(TIERS)))
    put(tk, tr, 13, f'=D{tr}*Efficiency*E{tr}*IF(C{tr}="firm",1,IF(C{tr}="household",{hh_},'
                    f'IF(C{tr}="account",{ac_},{nw_})))', INK, fmt=NUM)
    put(tk, tr, 14, f'=M{tr}/60*IF(K{tr}="operations",1+Switching_Uplift,1)', INK, fmt=HOURS)
    task_rows[t["id"]] = tr; tr += 1
TN = tr - 1
put(tk, tr + 1, 2, "Total operations hours a year", BOLD)
put(tk, tr + 1, 14, f'=SUMPRODUCT((K2:K{TN}="operations")*N2:N{TN})', BOLD, fmt=HOURS)
put(tk, tr + 3, 2, I["task_meta"]["seed_note"], MUTED, align="left")
tk.row_dimensions[tr + 3].height = 30
for nm, col in [("Tasks_Basis", "C"), ("Tasks_Min", "D"), ("Tasks_Occ", "E"),
                ("Tasks_A1", "F"), ("Tasks_A2", "G"), ("Tasks_A3", "H"), ("Tasks_A4", "I"),
                ("Tasks_Owner", "J"), ("Tasks_WT", "K"), ("Tasks_Tag", "L"), ("Tasks_Hours", "N")]:
    name(nm, "Tasks", f"${col}$2:${col}${TN}")

# ================================================================= ENGINE
en = sheet("Engine")
for i, w in enumerate([46, 17, 17, 17, 17, 66], 1): en.column_dimensions[gcl(i)].width = w
put(en, 1, 1, "ENGINE", TITLE, border=False)
put(en, 2, 1, "Every cell on this sheet is a formula. There are no typed numbers here — if one appears, "
              "it is a bug.", MUTED, border=False)
en.freeze_panes = "B4"
E = {}
er = 4
def sect(t):
    global er
    er += 1
    c = put(en, er, 1, t, HDR_F, fill=HDR_FILL); put(en, er, 2, "", INK, fill=HDR_FILL)
    for cc in range(3, 7): put(en, er, cc, "", INK, fill=HDR_FILL)
    er += 1
def val(key, label, formula, fmt=None, note=""):
    global er
    put(en, er, 1, label, INK, align="left")
    put(en, er, 2, formula, INK, fmt=fmt)
    if note: put(en, er, 6, note, MUTED, align="left")
    E[key] = er; er += 1

ADV_ROWS = [roster_rows[m["role_id"]] for m in I["roster"] if m["is_advisor"]]
sect("RATES")
val("blended", "Blended operations $ per hour",
    "=IFERROR(SUMPRODUCT(Roster_OpsHours,Roster_Hourly)/SUM(Roster_OpsHours),0)", MONEY2,
    "Weighted by the hours each operations role actually gives to operations.")
val("advhr", "Advisor loaded $ per hour",
    "=IFERROR(SUMPRODUCT(Roster_Adv,Roster_Active,Roster_Loaded)/(SUMPRODUCT(Roster_Adv,Roster_Active,Roster_FTE)*Productive_Hours),0)",
    MONEY2, "What an advisor hour COSTS. Used for case 4 on purpose.")
val("advrev", "Revenue per advisor hour",
    "=IFERROR(Revenue_Total/(SUMPRODUCT(Roster_Adv,Roster_Active,Roster_FTE)*Productive_Hours),0)", MONEY2,
    "What an advisor hour EARNS. Larger, and it invites the challenge that the hour would not have been sold.")

sect("BOOK AND REVENUE")
val("hh", "Total households", "=Total_HH", NUM)
val("aum", "Total AUM", "=AUM_Total", MONEY)
val("new", "New households per year", "=Total_HH*Growth_Rate", '#,##0.0')
val("rev", "Revenue at the fee schedule", "=Revenue_Total", MONEY)
val("revhh", "Revenue per household", "=IFERROR(Revenue_Total/Total_HH,0)", MONEY)
val("advcount", "Advisor count", "=SUMPRODUCT(Roster_Adv,Roster_Active)", NUM)
val("revadv", "Revenue per advisor", f"=IFERROR(Revenue_Total/B{E['advcount']},0)", MONEY)
val("revemp", "Revenue per employee (per FTE)", "=IFERROR(Revenue_Total/Total_FTE,0)", MONEY)

sect("OPERATIONS HOURS PER YEAR")
val("recur", "Recurring (household and account work)",
    f'=SUMPRODUCT((Tasks_WT="operations")*((Tasks_Basis="household")+(Tasks_Basis="account"))*Tasks_Hours)', HOURS)
val("onboard", "Onboarding (scales with growth, not the book)",
    f'=SUMPRODUCT((Tasks_WT="operations")*(Tasks_Basis="new_household")*Tasks_Hours)', HOURS)
val("firmh", "Firm-level (does not scale with households)",
    f'=SUMPRODUCT((Tasks_WT="operations")*(Tasks_Basis="firm")*Tasks_Hours)', HOURS)
val("req", "TOTAL REQUIRED", f"=B{E['recur']}+B{E['onboard']}+B{E['firmh']}", HOURS)
val("avail", "TOTAL AVAILABLE", "=Ops_Hours_Available", HOURS)
val("util", "UTILISATION", f"=IFERROR(B{E['req']}/B{E['avail']},0)", PCT,
    "Over 100% means there is more operations work than the operations team can hold.")

sect("WHERE THE OPERATIONS WORK SITS TODAY")
owner_ids = sorted({t["owner"] for t in I["tasks"]})
for rid in owner_ids:
    val("own_" + rid, f"Hours owned by {rid}",
        f'=SUMPRODUCT((Tasks_WT="operations")*(Tasks_Owner="{rid}")*Tasks_Hours)', HOURS)
adv_ids = [m["role_id"] for m in I["roster"] if m["is_advisor"]]
val("opsadv", "Operations work sitting on advisors",
    "=" + "+".join(f"B{E['own_' + r]}" for r in adv_ids if "own_" + r in E), HOURS,
    "Operations work an advisor currently does. This is what an operations seat moves.")

sect("PER TIER")
for i, t in enumerate(TIERS):
    put(en, er, 2 + i, t, HDR_F, fill=HDR_FILL)
put(en, er, 1, "", INK, fill=HDR_FILL); put(en, er, 6, "", INK, fill=HDR_FILL)
er += 1
ET = {}
def trow(key, label, f_of_i, fmt=None, note=""):
    global er
    put(en, er, 1, label, INK, align="left")
    for i in range(len(TIERS)):
        put(en, er, 2 + i, f_of_i(i), INK, fmt=fmt)
    if note: put(en, er, 6, note, MUTED, align="left")
    ET[key] = er; er += 1

A_COL = ["Tasks_A1", "Tasks_A2", "Tasks_A3", "Tasks_A4"]
bkrow = lambda i: BOOK_FIRST + i
trow("share", "Share of the book", lambda i: f"=Book!C{bkrow(i)}/Total_HH", PCT)
trow("recurmin", "Recurring operations minutes per household",
     lambda i: (f'=SUMPRODUCT((Tasks_WT="operations")*(Tasks_Basis="household")*{A_COL[i]}*Tasks_Min*Efficiency*Tasks_Occ)'
                f'+SUMPRODUCT((Tasks_WT="operations")*(Tasks_Basis="account")*{A_COL[i]}*Tasks_Min*Efficiency*Tasks_Occ)*Book!E{bkrow(i)}'), NUM)
trow("onbmin", "Onboarding operations minutes per new household",
     lambda i: f'=SUMPRODUCT((Tasks_WT="operations")*(Tasks_Basis="new_household")*{A_COL[i]}*Tasks_Min*Efficiency*Tasks_Occ)', NUM)
trow("advmin", "Advisory minutes per household",
     lambda i: (f'=SUMPRODUCT((Tasks_WT="advisory")*(Tasks_Basis="household")*{A_COL[i]}*Tasks_Min*Efficiency*Tasks_Occ)'
                f'+SUMPRODUCT((Tasks_WT="advisory")*(Tasks_Basis="account")*{A_COL[i]}*Tasks_Min*Efficiency*Tasks_Occ)*Book!E{bkrow(i)}'), NUM,
     "Advisory work never counts against operations capacity. It is here only for the fully loaded cost to serve.")
trow("dirh", "Direct operations hours per household",
     lambda i: f"={gcl(2+i)}{ET['recurmin']}/60*(1+Switching_Uplift)", HOURS)
trow("dircost", "Layer 1 — direct operations cost",
     lambda i: f"={gcl(2+i)}{ET['dirh']}*B{E['blended']}", MONEY)
trow("allocfirm", "Layer 2 — share of firm-level operations",
     lambda i: f"=IFERROR(B{E['firmh']}*B{E['blended']}/Total_HH,0)", MONEY)
trow("advcost", "Layer 3 — advisor time",
     lambda i: f"={gcl(2+i)}{ET['advmin']}/60*B{E['advhr']}", MONEY)
trow("ovh", "Layer 4 — non-payroll firm overhead", lambda i: "=Firm_Overhead_PH", MONEY)
trow("loaded", "Fully loaded cost to serve",
     lambda i: f"=SUM({gcl(2+i)}{ET['dircost']}:{gcl(2+i)}{ET['ovh']})", MONEY)
trow("revph", "Revenue per household", lambda i: f"=Book!L{bkrow(i)}", MONEY)
trow("mgloaded", "Margin, fully loaded",
     lambda i: f"={gcl(2+i)}{ET['revph']}-{gcl(2+i)}{ET['loaded']}", MONEY)
trow("mgdirect", "Margin on direct operations cost alone",
     lambda i: f"={gcl(2+i)}{ET['revph']}-{gcl(2+i)}{ET['dircost']}", MONEY)
trow("onbcost", "Onboarding cost per new household (one-off)",
     lambda i: f"={gcl(2+i)}{ET['onbmin']}/60*(1+Switching_Uplift)*B{E['blended']}", MONEY)

sect("CAPACITY AT TODAY'S TIER MIX (the hand-checkable closed form)")
val("fixed", "Fixed hours (firm-level only)", f"=B{E['firmh']}", HOURS)
ONB_MIX = "+".join(f"{gcl(2+i)}{ET['onbmin']}*Book!$F${BOOK_FIRST+i}" for i in range(len(TIERS)))
val("perhh", "Marginal hours per additional household",
    f"=(SUMPRODUCT(B{ET['share']}:{gcl(1+len(TIERS))}{ET['share']},B{ET['recurmin']}:{gcl(1+len(TIERS))}{ET['recurmin']})"
    f"+Growth_Rate*({ONB_MIX}))/60*(1+Switching_Uplift)", HOURS,
    "One more household at today's mix, including the onboarding work growth brings with it.")
val("breakhh", "Households at which capacity breaks",
    f"=IFERROR((B{E['avail']}-B{E['fixed']})/B{E['perhh']},\"\")", '#,##0.0')
val("breakyrs", "Years to reach that count",
    f'=IFERROR(IF(B{E["breakhh"]}<=Total_HH,0,LN(B{E["breakhh"]}/Total_HH)/LN(1+Growth_Rate)),"")', '#,##0.00')
val("breakdate", "Which is around",
    f'=IFERROR(As_Of+B{E["breakyrs"]}*365.25,"")', DATE,
    "Holds today's tier mix. Case2 projects the mix forward as well, which lands slightly earlier.")

# ============================================ ASSUMPTIONS: the rest of the ledger
n = 0
for m in I["roster"]:
    rid, rw = m["role_id"], roster_rows[m["role_id"]]
    for suffix, col, unit, tag, src, who, chk in [
        ("annual compensation", "G", "usd", m["comp_tag"], m["comp_src"], m["who"],
         "never" if m["comp_tag"] == "PLACEHOLDER" else I["meta"]["as_of"]),
        ("FTE", "C", "fte", m["fte_tag"],
         "Headcount I can see." if m["fte_tag"] == "OBSERVED" else "My estimate of the hours actually worked here.",
         "The principal holds the contract or the offer letter.", I["meta"]["as_of"])] + (
        [("share of time on operations", "F", "pct", "ESTIMATED",
          "My estimate of how much of this person's week is operations work rather than planning, bookkeeping or reception.",
          "The person themselves, from a two-week time log. The cheapest input in the model to upgrade.", I["meta"]["as_of"])]
        if m["is_operations"] else []):
        n += 1
        put(asm, lr, 1, f"R-{n}", MUTED)
        put(asm, lr, 2, f'=Roster!B{rw}&" — {suffix}"', LINK_F, align="left")
        put(asm, lr, 3, f"=Roster!{col}{rw}", LINK_F, fmt=UNIT_FMT.get(unit))
        put(asm, lr, 4, unit, MUTED)
        put(asm, lr, 5, tag, WARN_F if tag == "PLACEHOLDER" else INK); tag_dv.add(asm.cell(row=lr, column=5))
        put(asm, lr, 6, src, MUTED, align="left"); put(asm, lr, 7, who, MUTED, align="left")
        put(asm, lr, 8, chk, WARN_F if chk == "never" else MUTED); lr += 1
n = 0
for i, b in enumerate(I["book"]):
    n += 1; bw = BOOK_FIRST + i
    put(asm, lr, 1, f"B-{n}", MUTED)
    put(asm, lr, 2, f'="{b["tier"]} {b["label"]} — households, average AUM, accounts per household"', INK, align="left")
    put(asm, lr, 3, f'=Book!C{bw}&" households · "&TEXT(Book!D{bw},"$#,##0")&" avg · "&TEXT(Book!E{bw},"0.0")&" accts"', LINK_F)
    put(asm, lr, 4, "tier", MUTED)
    put(asm, lr, 5, b["tag"], INK); tag_dv.add(asm.cell(row=lr, column=5))
    put(asm, lr, 6, b["src"], MUTED, align="left"); put(asm, lr, 7, b["who"], MUTED, align="left")
    put(asm, lr, 8, I["meta"]["as_of"], MUTED); lr += 1
n += 1
put(asm, lr, 1, f"B-{n}", MUTED)
put(asm, lr, 2, "How new households distribute across the tiers", INK, align="left")
put(asm, lr, 3, f'=Book!A{BOOK_FIRST}&" "&TEXT(Book!F{BOOK_FIRST},"0%")&" · "&Book!A{BOOK_FIRST+1}&" "&TEXT(Book!F{BOOK_FIRST+1},"0%")&" · "&Book!A{BOOK_FIRST+2}&" "&TEXT(Book!F{BOOK_FIRST+2},"0%")&" · "&Book!A{BOOK_FIRST+3}&" "&TEXT(Book!F{BOOK_FIRST+3},"0%")', LINK_F)
put(asm, lr, 4, "mix", MUTED)
put(asm, lr, 5, I["book_meta"]["new_mix_tag"], INK); tag_dv.add(asm.cell(row=lr, column=5))
put(asm, lr, 6, I["book_meta"]["new_mix_src"], MUTED, align="left")
put(asm, lr, 7, I["book_meta"]["new_mix_who"], MUTED, align="left")
put(asm, lr, 8, I["meta"]["as_of"], MUTED); lr += 1
put(asm, lr, 1, "F-1", MUTED)
put(asm, lr, 2, f"Tiered fee schedule ({len(I['fee_schedule']['bands'])} bands)", INK, align="left")
put(asm, lr, 3, f'="see FeeSchedule tab — "&TEXT(FeeSchedule!D2,"0.00%")&" to "&TEXT(FeeSchedule!D{FN},"0.00%")', LINK_F)
put(asm, lr, 4, "schedule", MUTED)
put(asm, lr, 5, I["fee_schedule"]["tag"], INK); tag_dv.add(asm.cell(row=lr, column=5))
put(asm, lr, 6, I["fee_schedule"]["src"], MUTED, align="left")
put(asm, lr, 7, I["fee_schedule"]["who"], MUTED, align="left")
put(asm, lr, 8, I["meta"]["as_of"], MUTED); lr += 1
n = 0
for t in I["tasks"]:
    n += 1; tw = task_rows[t["id"]]
    put(asm, lr, 1, f"T-{n}", MUTED)
    put(asm, lr, 2, f'=Tasks!B{tw}&IF(SUM(Tasks!F{tw}:I{tw})=4,""," (selected tiers)")', LINK_F, align="left")
    put(asm, lr, 3, f'=TEXT(Tasks!D{tw},"#,##0")&" min x "&TEXT(Tasks!E{tw},"0.00")&"/yr · "&Tasks!C{tw}&" · "&Tasks!J{tw}', LINK_F)
    put(asm, lr, 4, "minutes", MUTED)
    put(asm, lr, 5, f"=Tasks!L{tw}", LINK_F)
    put(asm, lr, 6, f"Seeded from the shape of real RIA operations work. The minutes are not measured.", MUTED, align="left")
    put(asm, lr, 7, "Me, with a stopwatch. Three runs, median, interruptions included.", MUTED, align="left")
    put(asm, lr, 8, "never", WARN_F); lr += 1
LEDGER_LAST = lr - 1
LEDGER_FIRST = LEDGER_HDR + 1
put(asm, lr + 1, 2, "Inputs in the ledger", BOLD)
put(asm, lr + 1, 3, f"=COUNTA(B{LEDGER_FIRST}:B{LEDGER_LAST})", BOLD, fmt=NUM)
for i, tg in enumerate(["MEASURED", "OBSERVED", "BENCHMARK", "ESTIMATED", "PLACEHOLDER"]):
    put(asm, lr + 2 + i, 2, f"  of which {tg.lower()}", INK)
    put(asm, lr + 2 + i, 3, f'=COUNTIF(E{LEDGER_FIRST}:E{LEDGER_LAST},"{tg}")', WARN_F if tg == "PLACEHOLDER" else INK, fmt=NUM)
name("Ledger_Tags", "Assumptions", f"$E${LEDGER_FIRST}:$E${LEDGER_LAST}")
name("Ledger_Count", "Assumptions", f"$C${lr+1}")
name("Ledger_Placeholders", "Assumptions", f"$C${lr+6}")

# ================================================================= CASE 1
R7 = roster_rows["R7"]; R4 = roster_rows["R4"]
c1 = sheet("Case1_Departure")
for i, w in enumerate([44, 16, 16, 16, 16, 60], 1): c1.column_dimensions[gcl(i)].width = w
put(c1, 1, 1, "CASE 1 — WHAT HAPPENS IN 100 DAYS", TITLE, border=False)
put(c1, 2, 1, "The consultant's departure. The most urgent of the four and the only one with a date already on it.",
    MUTED, border=False)
c1.freeze_panes = "A4"
C1 = {}
r = 4
def c1v(key, label, f, fmt=None, note="", font=INK):
    global r
    put(c1, r, 1, label, INK, align="left"); put(c1, r, 2, f, font, fmt=fmt)
    if note: put(c1, r, 6, note, MUTED, align="left")
    C1[key] = r; r += 1
c1v("days", "Days until she leaves", "=Departure_Days", NUM)
c1v("date", "Which is", "=As_Of+Departure_Days", 'yyyy-mm-dd')
c1v("cat", "Hours a year the catalogue has captured for her", f"=Engine!B{E['own_R7']}", HOURS)
c1v("con", "Hours a year she is contracted for", f"=Roster!C{R7}*Productive_Hours", HOURS)
c1v("unm", f"Hours the catalogue has NOT captured", f"=MAX(0,B{C1['con']}-B{C1['cat']})", HOURS,
    "Almost certainly real work not yet written down, not slack. Everything below understates the problem by this much.",
    font=WARN_F)
r += 1
put(c1, r, 1, "BASE HOURS (read from the roster, independent of the scenario switch)", H2, border=False); r += 1
c1v("allops", "Operations hours available with her here",
    f"=SUMPRODUCT(Roster!E2:E{RN},Roster!C2:C{RN},Roster!F2:F{RN})*Productive_Hours+Extra_Ops_Fte*Productive_Hours", HOURS)
c1v("r7h", "  of which hers", f"=Roster!C{R7}*Roster!F{R7}*Productive_Hours", HOURS)
c1v("r4base", "  my operations hours today", f"=Roster!C{R4}*Roster!F{R4}*Productive_Hours", HOURS)
c1v("r4seat", "  my operations hours in the seat", f"=Roster!C{R4}*Seat_Ops_Alloc*Productive_Hours", HOURS)
c1v("req", "Operations hours required (unchanged by who does them)", f"=Engine!B{E['req']}", HOURS)
r += 1
put(c1, r, 1, "THE THREE PLACES HER HOURS CAN GO", H2, border=False); r += 1
OPT_HDR = r
hdr(c1, OPT_HDR, ["Response", "Annual cost", "Hours available after", "Utilisation after",
                  "Hours left with no owner", "What it costs elsewhere"])
c1.freeze_panes = c1.cell(row=4, column=1)
rows = [
    ("Absorb across the existing team", "=0", f"=B{C1['allops']}-B{C1['r7h']}",
     "No new spend. The hours land on the people already here."),
    ("Backfill with a like-for-like hire", "=Backfill_Cost", f"=B{C1['allops']}-B{C1['r7h']}+Roster!C{R7}*Productive_Hours",
     "Replaces her hours at the same cost. Break-even is immediate by construction; the question is whether the hours come back."),
    ("I take it on", f"=(Ops_Seat_Comp-Roster!G{R4})*(1+Benefits_Load)",
     f"=B{C1['allops']}-B{C1['r7h']}+B{C1['r4seat']}-B{C1['r4base']}",
     "Compensation delta only, loaded. Displaces paraplanning work that has to go somewhere — see Case4."),
]
r = OPT_HDR + 1
OPT_FIRST = r
for label, cost, avail, note in rows:
    put(c1, r, 1, label, INK, align="left")
    put(c1, r, 2, cost, INK, fmt=MONEY)
    put(c1, r, 3, avail, INK, fmt=HOURS)
    put(c1, r, 4, f"=IFERROR(B{C1['req']}/C{r},0)", INK, fmt=PCT)
    put(c1, r, 5, f"=MAX(0,B{C1['req']}-C{r})", INK, fmt=HOURS)
    put(c1, r, 6, note, MUTED, align="left")
    c1.row_dimensions[r].height = 28
    r += 1
OPT_LAST = r - 1
name("C1_Absorb_Util", "Case1_Departure", f"$D${OPT_FIRST}")
name("C1_Internal_Avail", "Case1_Departure", f"$C${OPT_LAST}")
name("C1_Internal_Cost", "Case1_Departure", f"$B${OPT_LAST}")
name("C1_Catalogued", "Case1_Departure", f"$B${C1['cat']}")
name("C1_Unmapped", "Case1_Departure", f"$B${C1['unm']}")
r += 1
put(c1, r, 1, "THE SENTENCE THIS ENDS ON", H2, border=False); r += 1
put(c1, r, 1, f'="When she leaves, "&TEXT(B{C1["cat"]},"#,##0")&" catalogued hours a year need a home — and "'
              f'&TEXT(B{C1["unm"]},"#,##0")&" more that this model cannot see yet. Absorbing them puts operations at "'
              f'&TEXT(D{OPT_FIRST},"0.0%")&" and leaves "&TEXT(E{OPT_FIRST},"#,##0")&" hours with nobody on them."', BOLD, align="left")
c1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); c1.row_dimensions[r].height = 30

# ================================================================= CASE 2
c2 = sheet("Case2_Capacity")
for i, w in enumerate([9, 13, 11, 11, 11, 11, 12, 12, 14, 14, 12, 8, 9], 1): c2.column_dimensions[gcl(i)].width = w
put(c2, 1, 1, "CASE 2 — WHERE CAPACITY BREAKS", TITLE, border=False)
put(c2, 2, 1, "Month by month. New households enter by the new-business mix, so the tier mix shifts as the book grows.",
    MUTED, border=False)
put(c2, 3, 1, "Monthly growth rate", BOLD); put(c2, 3, 2, "=(1+Growth_Rate)^(1/12)-1", INK, fmt='0.00%')
name("Monthly_Growth", "Case2_Capacity", "$B$3")
PROJ_HDR = 5
hdr(c2, PROJ_HDR, ["Month", "Date"] + [f"Households {t}" for t in TIERS] +
    ["Total households", "New / yr", "Hours required", "Hours available", "Utilisation", "Over?", "Over with seat?"])
c2.freeze_panes = c2.cell(row=PROJ_HDR + 1, column=3)
P0 = PROJ_HDR + 1
for m in range(HORIZON_MONTHS + 1):
    rr2 = P0 + m
    put(c2, rr2, 1, m, INK, fmt=NUM)
    put(c2, rr2, 2, f"=As_Of+ROUND(A{rr2}*30.4375,0)", INK, fmt='mmm yyyy')
    for i in range(len(TIERS)):
        col = gcl(3 + i)
        if m == 0:
            put(c2, rr2, 3 + i, f"=Book!C{BOOK_FIRST + i}", INK, fmt='#,##0.0')
        else:
            put(c2, rr2, 3 + i, f"={col}{rr2-1}+G{rr2-1}*Monthly_Growth*Book!$F${BOOK_FIRST + i}", INK, fmt='#,##0.0')
    put(c2, rr2, 7, f"=SUM(C{rr2}:F{rr2})", INK, fmt='#,##0.0')
    put(c2, rr2, 8, f"=G{rr2}*Growth_Rate", INK, fmt='#,##0.0')
    onb_mix = "+".join(f"Engine!${gcl(2+i)}${ET['onbmin']}*Book!$F${BOOK_FIRST+i}" for i in range(len(TIERS)))
    put(c2, rr2, 9,
        f"=(SUMPRODUCT(C{rr2}:F{rr2},Engine!$B${ET['recurmin']}:$E${ET['recurmin']})"
        f"+H{rr2}*({onb_mix}))/60*(1+Switching_Uplift)"
        f"+Engine!$B${E['firmh']}", INK, fmt=HOURS)
    put(c2, rr2, 10, f"=Engine!$B${E['avail']}", INK, fmt=HOURS)
    put(c2, rr2, 11, f"=IFERROR(I{rr2}/J{rr2},0)", INK, fmt=PCT)
    put(c2, rr2, 12, f"=IF(I{rr2}>=J{rr2},1,0)", INK, fmt=NUM)
    put(c2, rr2, 13, f"=IF(I{rr2}>=C1_Internal_Avail,1,0)", INK, fmt=NUM)
PL = P0 + HORIZON_MONTHS
name("Proj_Req", "Case2_Capacity", f"$I${P0}:$I${PL}")
name("Proj_HH", "Case2_Capacity", f"$G${P0}:$G${PL}")
name("Proj_Flag", "Case2_Capacity", f"$L${P0}:$L${PL}")
name("Proj_FlagSeat", "Case2_Capacity", f"$M${P0}:$M${PL}")
SUM0 = PL + 2
put(c2, SUM0, 1, "THE CROSSING", H2, border=False)
C2 = {}
r = SUM0 + 1
def c2v(key, label, f, fmt=None, note=""):
    global r
    put(c2, r, 1, label, INK, align="left"); c2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    put(c2, r, 5, f, INK, fmt=fmt)
    if note: put(c2, r, 7, note, MUTED, align="left")
    C2[key] = r; r += 1
c2v("row", "First month at or over capacity (position in the table)", '=IFERROR(MATCH(1,Proj_Flag,0),0)', NUM)
c2v("frac", "Crossing, in months from today",
    f'=IF(E{C2["row"]}=0,"",IF(E{C2["row"]}=1,0,(E{C2["row"]}-2)+(Engine!$B${E["avail"]}-INDEX(Proj_Req,E{C2["row"]}-1))'
    f'/(INDEX(Proj_Req,E{C2["row"]})-INDEX(Proj_Req,E{C2["row"]}-1))))', '#,##0.00')
c2v("hh", "Households at the crossing",
    f'=IF(E{C2["row"]}=0,"",IF(E{C2["row"]}=1,INDEX(Proj_HH,1),INDEX(Proj_HH,E{C2["row"]}-1)'
    f'+(E{C2["frac"]}-(E{C2["row"]}-2))*(INDEX(Proj_HH,E{C2["row"]})-INDEX(Proj_HH,E{C2["row"]}-1))))', '#,##0.0')
c2v("date", "Which is around", f'=IF(E{C2["row"]}=0,"",As_Of+ROUND(E{C2["frac"]}*30.4375,0))', DATE)
c2v("util", "Utilisation today", f"=Engine!B{E['util']}", PCT)
name("C2_BreakDate", "Case2_Capacity", f"$E${C2['date']}")
name("C2_BreakHH", "Case2_Capacity", f"$E${C2['hh']}")
name("C2_BreakMonths", "Case2_Capacity", f"$E${C2['frac']}")
r += 1
put(c2, r, 1, "SENSITIVITY OF THE DATE — change Growth_Rate on Assumptions and watch these three move together.",
    MUTED, align="left", border=False); r += 1
put(c2, r, 1, "Closed form at today's tier mix (hand-checkable)", INK, align="left")
c2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
put(c2, r, 5, f"=Engine!B{E['breakhh']}", INK, fmt='#,##0.0')
put(c2, r, 7, "Higher than the projection above, because the projection sends most new households into the larger "
              "tiers, which cost more hours each.", MUTED, align="left"); r += 1
put(c2, r, 1, "THE SENTENCE THIS ENDS ON", H2, border=False); r += 1
put(c2, r, 1, f'=IF(E{C2["row"]}=0,"At this growth rate operations does not run out of capacity inside the horizon.",'
              f'"At "&TEXT(Growth_Rate,"0%")&" growth, operations runs out of capacity at "&TEXT(E{C2["hh"]},"#,##0")'
              f'&" households, around "&TEXT(E{C2["date"]},"mmmm yyyy")&" — "&TEXT(E{C2["frac"]},"#,##0")&" months from now.")',
    BOLD, align="left")
c2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9); c2.row_dimensions[r].height = 28

# ================================================================= CASE 3
c3 = sheet("Case3_CostToServe")
for i, w in enumerate([46, 15, 15, 15, 15, 15, 58], 1): c3.column_dimensions[gcl(i)].width = w
put(c3, 1, 1, "CASE 3 — WHAT EACH CLIENT TIER COSTS TO SERVE", TITLE, border=False)
put(c3, 2, 1, "Presented as a question. This is not a finding about any client, and it is not a recommendation.",
    MUTED, border=False)
put(c3, 3, 1, "WARNING", WARN_F)
put(c3, 3, 2, f'="The largest layer below is non-payroll firm overhead at "&TEXT(Firm_Overhead_PH,"$#,##0")'
              f'&" per household. That number is a PLACEHOLDER — it was invented so the model would run. It decides '
              f'on its own whether the bottom tier looks profitable. Get the real figure before presenting this tab."',
    WARN_F, align="left")
c3.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7); c3.row_dimensions[3].height = 30
T3H = 5
hdr(c3, T3H, ["Per household per year"] + TIERS + ["Whole book", "Note"])
c3.freeze_panes = c3.cell(row=T3H + 1, column=2)
BOOKCOL = gcl(2 + len(TIERS))
r = T3H + 1
def c3row(label, eng_row, fmt=MONEY, note="", weighted=True, bold=False):
    global r
    f = BOLD if bold else INK
    put(c3, r, 1, label, f, align="left")
    for i in range(len(TIERS)):
        put(c3, r, 2 + i, f"=Engine!{gcl(2+i)}{eng_row}", f, fmt=fmt)
    if weighted:
        terms = "+".join(f"{gcl(2+i)}{r}*Book!C{BOOK_FIRST+i}" for i in range(len(TIERS)))
        put(c3, r, 2 + len(TIERS), f"=IFERROR(({terms})/Total_HH,0)", f, fmt=fmt)
    if note: put(c3, r, 3 + len(TIERS), note, MUTED, align="left")
    r += 1
c3row("Households", None, NUM, weighted=False)
for i in range(len(TIERS)):
    c3.cell(row=r - 1, column=2 + i).value = f"=Book!C{BOOK_FIRST+i}"
c3.cell(row=r - 1, column=2 + len(TIERS)).value = "=Total_HH"
c3.cell(row=r - 1, column=2 + len(TIERS)).number_format = NUM
c3row("Average AUM", None, MONEY, weighted=False)
for i in range(len(TIERS)):
    c3.cell(row=r - 1, column=2 + i).value = f"=Book!D{BOOK_FIRST+i}"
c3.cell(row=r - 1, column=2 + len(TIERS)).value = "=IFERROR(AUM_Total/Total_HH,0)"
c3.cell(row=r - 1, column=2 + len(TIERS)).number_format = MONEY
c3row("Revenue per household", ET["revph"], MONEY, "At the tiered schedule, after fee realisation.")
c3row("Direct operations hours", ET["dirh"], HOURS, "From the task catalogue. The only layer built from timings.")
c3row("Layer 1 — direct operations cost", ET["dircost"], MONEY)
c3row("Layer 2 — share of firm-level operations", ET["allocfirm"], MONEY, "Firm-level operations hours spread evenly.")
c3row("Layer 3 — advisor time", ET["advcost"], MONEY, "Advisory work at what an advisor hour costs.")
c3row("Layer 4 — non-payroll firm overhead", ET["ovh"], MONEY, "PLACEHOLDER. This is the layer that decides the answer.")
c3row("Fully loaded cost to serve", ET["loaded"], MONEY, "", bold=True)
c3row("Margin, fully loaded", ET["mgloaded"], MONEY, "", bold=True)
c3row("Margin on direct operations cost alone", ET["mgdirect"], MONEY,
      "Positive here and negative above means the gap is created by overhead and advisor time, not by operations.")
c3row("Onboarding cost per new household (one-off)", ET["onbcost"], MONEY,
      "Deliberately excluded from cost to serve. Smearing it across the settled book would overstate every tier.")
MG_ROW = T3H + 10
r += 1
put(c3, r, 1, "THE THREE VERSIONS OF THIS DECISION — laid out flat on purpose. The model has no view on which is right.", H2, border=False)
r += 1
for i, (t, s) in enumerate([
    ("Raise the minimum",
     "Stop taking households below a threshold. Protects capacity and margin. Costs the firm its pipeline: a small household today is sometimes a large one in ten years, and referrals come from families, not balances."),
    ("Serve that tier differently",
     "Keep them, change the service model: fewer scheduled meetings, group or digital review, a defined service tier. Cuts the operations hours each household takes. Costs consistency, and someone has to tell them."),
    ("Accept it as a pipeline cost",
     "Decide the shortfall is what the firm pays for relationships, referrals and the next generation. Perfectly defensible — but it should be a decision somebody made, not an accident of never having counted.")], 1):
    put(c3, r, 1, f"{i}. {t}", BOLD, align="left")
    put(c3, r, 2, s, MUTED, align="left")
    c3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7); c3.row_dimensions[r].height = 30; r += 1
r += 1
put(c3, r, 1, "THE SENTENCE THIS ENDS ON", H2, border=False); r += 1
put(c3, r, 1, f'="The bottom tier costs roughly "&TEXT(B{T3H+9},"$#,##0")&" a year to serve on a fully loaded basis '
              f'and produces "&TEXT(B{T3H+3},"$#,##0")&". That is a decision, not a finding, and there are three versions of it above."',
    BOLD, align="left")
c3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7); c3.row_dimensions[r].height = 28

# ================================================================= CASE 4
c4 = sheet("Case4_Seat")
for i, w in enumerate([48, 16, 16, 16, 16, 60], 1): c4.column_dimensions[gcl(i)].width = w
put(c4, 1, 1, "CASE 4 — THE OPERATIONS SEAT", TITLE, border=False)
put(c4, 2, 1, "An operations seat does not generate revenue. It protects advisor capacity and prevents service "
              "failures. That is the whole claim.", MUTED, border=False)
c4.freeze_panes = "A4"
C4 = {}
r = 4
def c4v(key, label, f, fmt=None, note="", font=INK):
    global r
    put(c4, r, 1, label, INK, align="left"); put(c4, r, 2, f, font, fmt=fmt)
    if note: put(c4, r, 6, note, MUTED, align="left")
    C4[key] = r; r += 1
c4v("delta", "Compensation delta for the seat", f"=Ops_Seat_Comp-Roster!G{R4}", MONEY)
c4v("cost", "Cost of the seat, loaded", f"=B{r-1}*(1+Benefits_Load)", MONEY)
name("C4_SeatCost", "Case4_Seat", f"$B${C4['cost']}")
r += 1
put(c4, r, 1, "BLENDED OPERATIONS RATE AFTER SHE GOES AND I TAKE THE SEAT", H2, border=False); r += 1
BR_HDR = r
hdr(c4, BR_HDR, ["Operations role after the change", "Hours", "Loaded $/hour", "Hours x rate"])
r = BR_HDR + 1
BR0 = r
ops_after = [m for m in I["roster"] if m["is_operations"] and m["role_id"] != "R7"]
for m in ops_after:
    rw = roster_rows[m["role_id"]]
    is_seat = m["role_id"] == "R4"
    put(c4, r, 1, f"=Roster!B{rw}", LINK_F, align="left")
    put(c4, r, 2, f"=Roster!C{rw}*{'Seat_Ops_Alloc' if is_seat else f'Roster!F{rw}'}*Productive_Hours", INK, fmt=HOURS)
    comp = "Ops_Seat_Comp" if is_seat else f"Roster!G{rw}"
    put(c4, r, 3, f"=IFERROR({comp}*(1+IF(Roster!I{rw}=1,0,Benefits_Load))/(Roster!C{rw}*Productive_Hours),0)", INK, fmt=MONEY2)
    put(c4, r, 4, f"=B{r}*C{r}", INK, fmt=MONEY)
    r += 1
BR1 = r - 1
put(c4, r, 1, "Blended rate after the change", BOLD)
put(c4, r, 2, f"=SUM(B{BR0}:B{BR1})", BOLD, fmt=HOURS)
put(c4, r, 3, f"=IFERROR(SUM(D{BR0}:D{BR1})/SUM(B{BR0}:B{BR1}),0)", BOLD, fmt=MONEY2)
BLEND_AFTER = f"C{r}"
r += 2
put(c4, r, 1, "WHAT THE SEAT PROTECTS", H2, border=False); r += 1
PR_HDR = r
hdr(c4, PR_HDR, ["Component", "Hours", "Valued at", "Value per year", "", "Note"])
r = PR_HDR + 1
PR0 = r
put(c4, r, 1, "Operations hours she leaves behind, covered in-house", INK, align="left")
put(c4, r, 2, "=C1_Catalogued", INK, fmt=HOURS)
put(c4, r, 3, f"={BLEND_AFTER}", INK, fmt=MONEY2)
put(c4, r, 4, f"=B{r}*C{r}", INK, fmt=MONEY); r += 1
put(c4, r, 1, "Operations work currently sitting on advisors, moved off", INK, align="left")
put(c4, r, 2, f"=Engine!B{E['opsadv']}", INK, fmt=HOURS)
put(c4, r, 3, f"=Engine!B{E['advhr']}", INK, fmt=MONEY2)
put(c4, r, 4, f"=B{r}*C{r}", INK, fmt=MONEY)
put(c4, r, 6, "Valued at what an advisor hour COSTS, not what it earns. The larger number invites the challenge "
              "that the hour would not have been sold.", MUTED, align="left"); r += 1
put(c4, r, 1, "Service failures avoided", INK, align="left")
put(c4, r, 2, "=Failures_Avoided", INK, fmt=NUM)
put(c4, r, 3, "=Failure_Cost", INK, fmt=MONEY)
put(c4, r, 4, f"=B{r}*C{r}", INK, fmt=MONEY)
put(c4, r, 6, "Deliberately zero. No incident list, so nothing is counted. Raise it only against incidents "
              "somebody can name.", MUTED, align="left"); r += 1
PR1 = r - 1
put(c4, r, 1, "Total protected", BOLD); put(c4, r, 4, f"=SUM(D{PR0}:D{PR1})", BOLD, fmt=MONEY)
PROT = f"D{r}"; r += 1
put(c4, r, 1, "Cost of the seat, loaded", INK); put(c4, r, 4, "=C4_SeatCost", INK, fmt=MONEY); r += 1
put(c4, r, 1, "Surplus", BOLD); put(c4, r, 4, f"={PROT}-C4_SeatCost", BOLD, fmt=MONEY)
SURPLUS = f"D{r}"; r += 1
put(c4, r, 1, "Cover ratio", BOLD); put(c4, r, 4, f"=IFERROR({PROT}/C4_SeatCost,0)", BOLD, fmt='0.00"x"')
r += 2
put(c4, r, 1, "WHAT IT COSTS ELSEWHERE — the hole in my own proposal", H2, border=False); r += 1
put(c4, r, 1, "Paraplanning hours displaced by the seat", INK, align="left")
put(c4, r, 2, f"=(Seat_Ops_Alloc-Roster!F{R4})*Roster!C{R4}*Productive_Hours", WARN_F, fmt=HOURS)
DISP = f"B{r}"
put(c4, r, 6, "The seat frees this many operations hours by moving my week — and displaces exactly the same number "
              "of paraplanning hours. That work does not disappear.", MUTED, align="left"); r += 1
put(c4, r, 1, "  if advisors absorb it", INK, align="left")
put(c4, r, 2, f"={DISP}*Engine!B{E['advhr']}", WARN_F, fmt=MONEY)
put(c4, r, 6, "At this cost the proposal is worse than doing nothing. This is the question to put to him.", MUTED, align="left"); r += 1
put(c4, r, 1, "  if a junior seat absorbs it", INK, align="left")
put(c4, r, 2, f"=IFERROR(Para_Backfill_Comp*(1+Benefits_Load)*{DISP}/Productive_Hours,0)", INK, fmt=MONEY); r += 1
put(c4, r, 1, "  if the firm decides some of it stops", INK, align="left")
put(c4, r, 2, "free, but somebody has to choose", MUTED); r += 1
put(c4, r, 1, "Capacity headroom the seat buys", INK, align="left")
put(c4, r, 2, '=IFERROR(MATCH(1,Proj_FlagSeat,0)-1,"beyond the horizon")', INK, fmt=NUM)
put(c4, r, 3, "months", MUTED); r += 2
put(c4, r, 1, "THE SENTENCE THIS ENDS ON", H2, border=False); r += 1
put(c4, r, 1, f'="The seat costs "&TEXT(C4_SeatCost,"$#,##0")&" loaded. It has to protect "&TEXT(C4_SeatCost,"$#,##0")'
              f'&" to pay for itself. On these numbers it protects "&TEXT({PROT},"$#,##0")&" — but only if the "'
              f'&TEXT({DISP},"#,##0")&" hours of planning work it displaces are not simply handed back to the advisors."',
    BOLD, align="left")
c4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); c4.row_dimensions[r].height = 30

# ================================================================= DASHBOARD
db = sheet("Dashboard")
for i, w in enumerate([44, 20, 16, 16, 16, 16], 1): db.column_dimensions[gcl(i)].width = w
put(db, 1, 1, "OPERATIONS CAPACITY & ECONOMICS", TITLE, border=False)
put(db, 2, 1, f'="Working model · as of "&TEXT(As_Of,"d mmmm yyyy")&" · scenario: "&Assumptions!C6', MUTED, border=False)
put(db, 3, 1, '="Of "&Ledger_Count&" inputs, "&Ledger_Placeholders&" are still placeholders — numbers invented so '
              'the model would run. Treat any figure resting on one as a structure, not a finding."', WARN_F, border=False)
db.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
r = 5
def dash(label, f, fmt=None, warn=False):
    global r
    put(db, r, 1, label, INK, align="left")
    put(db, r, 2, f, WARN_F if warn else BOLD, fmt=fmt); r += 1
put(db, r, 1, "THE BOOK", H2, border=False); r += 1
dash("Households", "=Total_HH", NUM)
dash("Assets under management", "=AUM_Total", MONEY)
dash("Revenue at the fee schedule", "=Revenue_Total", MONEY)
dash("Revenue per household", f"=Engine!B{E['revhh']}", MONEY)
dash("Revenue per advisor", f"=Engine!B{E['revadv']}", MONEY)
dash("Revenue per employee", f"=Engine!B{E['revemp']}", MONEY)
r += 1
put(db, r, 1, "OPERATIONS TODAY", H2, border=False); r += 1
dash("Hours required per year", f"=Engine!B{E['req']}", HOURS)
dash("Hours available per year", f"=Engine!B{E['avail']}", HOURS)
dash("Utilisation", f"=Engine!B{E['util']}", PCT, warn=True)
dash("Operations work sitting on advisors", f"=Engine!B{E['opsadv']}", HOURS)
r += 1
put(db, r, 1, "THE FOUR CASES", H2, border=False); r += 1
dash("1 · Catalogued hours her departure leaves behind", "=C1_Catalogued", HOURS)
dash("1 · Hours it has not captured", "=C1_Unmapped", HOURS, warn=True)
dash("1 · Utilisation if the team absorbs them", "=C1_Absorb_Util", PCT, warn=True)
dash("2 · Households at the capacity crossing", "=C2_BreakHH", '#,##0.0')
dash("2 · Which is around", "=C2_BreakDate", DATE, warn=True)
dash("3 · Bottom tier, fully loaded cost to serve", f"=Case3_CostToServe!B{T3H+9}", MONEY)
dash("3 · Bottom tier, revenue", f"=Case3_CostToServe!B{T3H+3}", MONEY)
dash("3 · Bottom tier, margin", f"=Case3_CostToServe!B{T3H+10}", MONEY, warn=True)
dash("4 · Cost of the operations seat, loaded", "=C4_SeatCost", MONEY)
dash("4 · What it protects", f"=Case4_Seat!{PROT}", MONEY)
dash("4 · Surplus", f"=Case4_Seat!{SURPLUS}", MONEY)

CH0 = r + 2
put(db, CH0, 1, "TIER ECONOMICS (chart source)", H2, border=False)
hdr(db, CH0 + 1, ["Tier", "Revenue per household", "Fully loaded cost to serve"])
for i, t in enumerate(TIERS):
    rr3 = CH0 + 2 + i
    put(db, rr3, 1, f"=Book!A{BOOK_FIRST+i}&\" \"&Book!B{BOOK_FIRST+i}", LINK_F)
    put(db, rr3, 2, f"=Engine!{gcl(2+i)}{ET['revph']}", INK, fmt=MONEY)
    put(db, rr3, 3, f"=Engine!{gcl(2+i)}{ET['loaded']}", INK, fmt=MONEY)
TIER_FIRST, TIER_LAST = CH0 + 2, CH0 + 1 + len(TIERS)

lc = LineChart(); lc.title = "Operations hours required vs available"
lc.y_axis.title = "Hours per year"; lc.x_axis.title = "Month"
lc.height, lc.width = 9.5, 21
data = Reference(c2, min_col=9, max_col=10, min_row=PROJ_HDR, max_row=PL)
cats = Reference(c2, min_col=2, min_row=P0, max_row=PL)
lc.add_data(data, titles_from_data=True); lc.set_categories(cats)
lc.series[0].graphicalProperties.line.width = 22000
lc.series[1].graphicalProperties.line.width = 22000
for s in lc.series: s.smooth = False
db.add_chart(lc, f"E{CH0 - 22}")

bc = BarChart(); bc.type = "col"; bc.title = "Revenue against fully loaded cost to serve, by tier"
bc.y_axis.title = "Dollars per household per year"; bc.x_axis.title = "Tier"
bc.height, bc.width = 9.5, 21
bdata = Reference(db, min_col=2, max_col=3, min_row=CH0 + 1, max_row=TIER_LAST)
bcats = Reference(db, min_col=1, min_row=TIER_FIRST, max_row=TIER_LAST)
bc.add_data(bdata, titles_from_data=True); bc.set_categories(bcats)
bc.gapWidth = 60
db.add_chart(bc, f"E{CH0 - 1}")

# openpyxl writes formulas with no cached values, so tell Excel and Numbers to
# compute everything the moment the file opens. Without this the sheet looks
# blank until someone forces a recalculation.
wb.calculation.fullCalcOnLoad = True

# --- cell map, so the parity check targets exact addresses -----------------
cellmap = {
  "engine": {k: f"B{v}" for k, v in E.items()},
  "engine_tier": {k: [f"{gcl(2+i)}{v}" for i in range(len(TIERS))] for k, v in ET.items()},
  "case1": {k: f"B{v}" for k, v in C1.items()},
  "case1_options": {"first_row": OPT_FIRST, "last_row": OPT_LAST,
                    "cost": "B", "avail": "C", "util": "D", "short": "E"},
  "case2": {k: f"E{v}" for k, v in C2.items()},
  "case3_header_row": T3H,
  "case4": {k: f"B{v}" for k, v in C4.items()},
  "case4_cells": {"blend_after": BLEND_AFTER, "protect_total": PROT, "surplus": SURPLUS,
                  "protect_first": PR0, "protect_last": PR1, "displaced": DISP},
  "fee_check_first": HC + 3, "fee_check_last": HC + 5,
  "ledger": {"first": LEDGER_FIRST, "last": LEDGER_LAST},
  "roster_rows": roster_rows, "task_rows": task_rows, "book_first": BOOK_FIRST,
}
json.dump(cellmap, open("dist/cellmap.json", "w"), indent=1)

wb.save("dist/RIA_Operations_Model.xlsx")
print("saved dist/RIA_Operations_Model.xlsx")
